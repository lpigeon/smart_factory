import openpyxl
from datetime import datetime
import time
import serial.tools.list_ports
import threading


# Arduino Uno를 찾아서 시리얼 포트에 연결합니다.
def connect_to_arduino_uno():
    ports = serial.tools.list_ports.comports()
    for port in ports:
        if "IOUSBHostDevice" in port.description:
            try:
                ser = serial.Serial(port.device, baudrate=9600)
                return ser
            except serial.SerialException:
                pass
    return None

serial_receive_data = ""
# 시리얼통신 수신 쓰레드 함수
def serial_read_thread():
    global serial_receive_data
    while True:
        read_data = ser.readline().decode()
        serial_receive_data = read_data

#컨베이어벨트 제어
def send_conveyor_speed(speed):
    if 0 <= speed <=255:
        ser.write(f"CV_MOTOR={speed}\n".encode())
    else:
        print("0~255사이의 값을 입력하세요")

def send_lamp_red(on_off):
    if on_off:
        ser.write("LAMP_RED=ON\n".encode())
    else:
        ser.write("LAMP_RED=OFF\n".encode())

def send_lamp_yellow(on_off):
    if on_off:
        ser.write("LAMP_YELLOW=ON\n".encode())
    else:
        ser.write("LAMP_YELLOW=OFF\n".encode())

def send_lamp_green(on_off):
    if on_off:
        ser.write("LAMP_GREEN=ON\n".encode())
    else:
        ser.write("LAMP_GREEN=OFF\n".encode())

# Arduino Uno와 연결합니다.
ser = connect_to_arduino_uno()

# 쓰레드를 시작합니다.
t1 = threading.Thread(target=serial_read_thread)
t1.daemon = True
t1.start()

# 엑셀 파일 열기 또는 생성
try:
    workbook = openpyxl.load_workbook('데이터저장.xlsx')
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save('데이터저장.xlsx')  # 새로운 엑셀 파일 생성

sheet = workbook.active

time.sleep(2.0)
print("start")
serial_receive_data = ""
row = sheet.max_row + 1
try:
    while True:
        # 투입쪽에 물건이 들어오면 컨베이어 동작
        if "PS_3=ON" in serial_receive_data:
            send_conveyor_speed(255)
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S") 
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "입구센서 ON"
            row += 1
            workbook.save('데이터저장.xlsx') 
        elif "PS_3=OFF" in serial_receive_data:
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")  
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "입구센서 OFF"
            row += 1
            workbook.save('데이터저장.xlsx') 
        elif "PS_2=ON" in serial_receive_data:
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")  
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "중앙센서 ON"
            row += 1
            workbook.save('데이터저장.xlsx') 
        elif "PS_2=OFF" in serial_receive_data:
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")  
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "중앙센서 OFF"
            row += 1
            workbook.save('데이터저장.xlsx') 
        elif "PS_1=ON" in serial_receive_data:
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")  
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "출구센서 ON"
            row += 1
            workbook.save('데이터저장.xlsx') 
        # 출구쪽센서에 물건이 들어오면 컨베이어 멈춤
        elif "PS_1=OFF" in serial_receive_data:
            print(serial_receive_data)
            serial_receive_data = ""
            now = datetime.now()
            current_time = now.strftime("%Y-%m-%d %H:%M:%S")  
            sheet.cell(row=row, column=1).value = current_time
            sheet.cell(row=row, column=2).value = "출구센서 OFF"
            row += 1
            workbook.save('데이터저장.xlsx') 
            time.sleep(2.0)
            send_conveyor_speed(0)
            send_lamp_red(False)
            send_lamp_yellow(False)
            send_lamp_green(True)

except KeyboardInterrupt:
    send_conveyor_speed(0)
    send_lamp_red(False)
    send_lamp_yellow(False)
    send_lamp_green(False)
    workbook.close()
    ser.close()