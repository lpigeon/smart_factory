import openpyxl
from datetime import datetime
import time

# 엑셀 파일 열기 또는 생성
try:
    workbook = openpyxl.load_workbook('시간저장.xlsx')
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.save('시간저장.xlsx')  # 새로운 엑셀 파일 생성

sheet = workbook.active

try:
    # 현재 시간을 아래 셀에 1초마다 저장
    row = sheet.max_row + 1  # 다음 행에 저장하기 위해 현재 마지막 행 번호 + 1로 설정
    while True:
        now = datetime.now()
        current_time = now.strftime("%Y-%m-%d %H:%M:%S")  # 시간 형식을 원하는 대로 수정할 수 있습니다
        print(current_time)
        sheet.cell(row=row, column=1).value = current_time
        row += 1
        workbook.save('시간저장.xlsx')  # 엑셀 파일 저장
        time.sleep(1)
except KeyboardInterrupt:
    workbook.close()
